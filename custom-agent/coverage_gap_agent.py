"""
Test Coverage Gap Report Agent
Cross-references all Jira stories vs TestRail test cases,
identifies stories with no test coverage, and exports a traceability
matrix to a styled Excel workbook.
"""

import os
import sys
import json
from datetime import datetime

from dotenv import load_dotenv

# Load .env from the rag directory (shared config)
load_dotenv(os.path.join(os.path.dirname(__file__), "../rag/.env"))

# Make the rag package importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../rag"))

from langchain.agents import create_agent
from langchain_core.tools import tool
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Environment helpers
# ---------------------------------------------------------------------------

PROJECT_NAME = os.getenv("PROJECT_NAME", "KAN")

# ---------------------------------------------------------------------------
# Tool 1: RAG Retriever
# ---------------------------------------------------------------------------

@tool
def rag_retriever(query: str, top_k: int = 10) -> str:
    """
    Query the RAG vector store to retrieve relevant indexed documents.
    Use this to look up Jira stories, TestRail test cases, API contracts, and schema.

    Examples:
      - query="all Jira stories KAN project" to get the full list of stories
      - query="<story summary>" to find TestRail test cases matching a story

    Returns a JSON array of matching documents with their id, content, and metadata.
    """
    from get_relevant_docs import get_relevant_docs  # lazy import after sys.path is set

    try:
        docs = get_relevant_docs(query, top_k=top_k)
        return json.dumps(docs, indent=2)
    except Exception as e:
        return json.dumps({"error": str(e)})


# ---------------------------------------------------------------------------
# Tool 2: Export Gap Report to Excel
# ---------------------------------------------------------------------------

@tool
def export_gap_report_to_excel(coverage_data_json: str) -> str:
    """
    Export the test coverage gap analysis to a styled Excel workbook.

    coverage_data_json must be a JSON array of objects with keys:
      - jira_key      (str)  e.g. "KAN-6"
      - summary       (str)
      - status        (str)
      - issuetype     (str)
      - test_cases    (list of str) – matched TC IDs like ["C59","C61"]
      - coverage      (str) – "Covered", "Partial", or "No Coverage"

    Returns the path of the saved Excel file.
    """
    output_dir = os.path.join(os.path.dirname(__file__), "reports")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"coverage_gap_report_{timestamp}.xlsx")

    try:
        coverage_data = json.loads(coverage_data_json)
    except json.JSONDecodeError as e:
        return f"ERROR: Invalid JSON — {e}"

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Detailed Traceability Matrix
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Coverage Gap Report"

    # Colour palette
    header_fill = PatternFill("solid", fgColor="1F3864")   # dark navy
    covered_fill = PatternFill("solid", fgColor="C6EFCE")  # green
    partial_fill  = PatternFill("solid", fgColor="FFEB9C") # yellow/orange
    gap_fill      = PatternFill("solid", fgColor="FFC7CE") # red
    alt_row_fill  = PatternFill("solid", fgColor="F2F2F2") # light grey

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    title_font   = Font(bold=True, size=14, color="1F3864")
    bold_font    = Font(bold=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_w  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = "Test Coverage Gap Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Sub-title (generation timestamp)
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ["Jira Key", "Summary", "Issue Type", "Status", "Test Cases", "Coverage"]
    header_row = 3
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # Data rows
    for row_idx, item in enumerate(coverage_data, start=header_row + 1):
        coverage = item.get("coverage", "No Coverage")
        row_fill = covered_fill if coverage == "Covered" else (partial_fill if coverage == "Partial" else gap_fill)
        alt = alt_row_fill if row_idx % 2 == 0 else None

        values = [
            item.get("jira_key", ""),
            item.get("summary", ""),
            item.get("issuetype", ""),
            item.get("status", ""),
            ", ".join(item.get("test_cases", [])) or "—",
            coverage,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = left_w if col == 2 else center
            # Apply coverage colour to coverage column; alternate fill elsewhere
            if col == 6:
                cell.fill = row_fill
                cell.font = bold_font
            elif alt:
                cell.fill = alt
        ws.row_dimensions[row_idx].height = 20

    # Column widths
    col_widths = [12, 45, 14, 16, 28, 16]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze panes below header
    ws.freeze_panes = "A4"

    # ------------------------------------------------------------------ #
    # Sheet 2: Summary Dashboard
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("Summary")

    covered_count = sum(1 for r in coverage_data if r.get("coverage") == "Covered")
    partial_count  = sum(1 for r in coverage_data if r.get("coverage") == "Partial")
    gap_count      = sum(1 for r in coverage_data if r.get("coverage") == "No Coverage")
    total          = len(coverage_data)
    coverage_pct   = round((covered_count / total * 100) if total else 0, 1)

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Coverage Summary"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 30

    summary_rows = [
        ("Metric", "Count", "Percentage"),
        ("Total Jira Stories",        total,          "100%"),
        ("Covered",                   covered_count,  f"{coverage_pct}%"),
        ("Partial Coverage",          partial_count,  f"{round(partial_count/total*100,1) if total else 0}%"),
        ("No Coverage (Gap)",         gap_count,      f"{round(gap_count/total*100,1) if total else 0}%"),
    ]

    for r_idx, row_data in enumerate(summary_rows, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = center
            if r_idx == 2:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx == 3:
                cell.fill = covered_fill
            elif r_idx == 4:
                cell.fill = partial_fill
            elif r_idx == 5:
                cell.fill = gap_fill
                cell.font = bold_font
        ws2.row_dimensions[r_idx].height = 22

    # Gaps list
    ws2["A8"] = "Stories with No Test Coverage:"
    ws2["A8"].font = bold_font
    gap_stories = [r for r in coverage_data if r.get("coverage") == "No Coverage"]
    for i, story in enumerate(gap_stories, start=9):
        ws2.cell(row=i, column=1, value=story.get("jira_key", ""))
        ws2.cell(row=i, column=2, value=story.get("summary", "")).alignment = left_w
        ws2.cell(row=i, column=1).fill = gap_fill

    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 20 if col > 1 else 18

    wb.save(output_path)
    return f"SUCCESS: Report saved to {output_path}"


# ---------------------------------------------------------------------------
# Agent definition
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are a QA Traceability Agent specializing in test coverage analysis.
You have exactly TWO tools: rag_retriever and export_gap_report_to_excel.

Your mission: Generate a complete Test Coverage Gap Report using only the RAG vector store.

Follow these steps EXACTLY in order:

Step 1 — Retrieve all Jira stories:
  Call rag_retriever with query="all Jira KAN project stories tasks epics" and top_k=30.
  Parse the returned documents to extract every unique Jira issue (look for documents whose
  id starts with "jira_" or whose content contains "KAN-").
  Collect: jira_key, summary, issuetype, status for each.

Step 2 — For EVERY Jira story found in Step 1:
  Call rag_retriever with the story summary as the query and top_k=8.
  From the returned documents, filter those whose id starts with "testrail_".
  Extract their test case IDs (e.g. "testrail_59" → "C59").

Step 3 — Classify each story:
  - "Covered"     → 2 or more matching TestRail cases found
  - "Partial"     → exactly 1 matching TestRail case found
  - "No Coverage" → 0 matching TestRail cases found

Step 4 — Build a JSON array where each element is:
  {{
    "jira_key":   "<KAN-XX>",
    "summary":    "<story summary>",
    "issuetype":  "<Story|Epic|Task>",
    "status":     "<Jira status>",
    "test_cases": ["C59", "C61", ...],
    "coverage":   "Covered|Partial|No Coverage"
  }}

Step 5 — Call export_gap_report_to_excel with the JSON string of that array.
  You MUST call this tool — do not describe what you would do, actually call it.

Step 6 — Report back:
  - The exact file path returned by export_gap_report_to_excel
  - Total stories, covered count, partial count, gap count
  - Bulleted list of stories with No Coverage

Be thorough. Process ALL stories found — do not skip any.
"""

def build_agent():
    tools = [rag_retriever, export_gap_report_to_excel]
    return create_agent(
        model="openai:gpt-4o",
        tools=tools,
        system_prompt=SYSTEM_PROMPT,
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=" * 65)
    print("  Test Coverage Gap Report Agent")
    print("=" * 65)

    agent = build_agent()
    user_prompt = (
        f"Generate a complete test coverage gap report for the '{PROJECT_NAME}' Jira project. "
        "Use the RAG retriever to retrieve all Jira stories and then find matching TestRail test cases "
        "for each story. Classify coverage, then call export_gap_report_to_excel to save the "
        "traceability matrix to Excel."
    )
    result = agent.invoke({"messages": [("human", user_prompt)]})

    print("\n" + "=" * 65)
    print("AGENT OUTPUT:")
    print("=" * 65)
    final_message = result["messages"][-1]
    print(getattr(final_message, "content", str(final_message)))
